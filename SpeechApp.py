import sys
from PyQt6.QtWidgets import QApplication, QWidget,QPushButton
from PyQt6.QtCore import QTimer
import speech_recognition as sr
from openpyxl import Workbook, load_workbook
from speechtotext import SpeechToText
from circlebutton import CircleButton
from timerthread import TimerThread
from listenthread import ListenThread
from commandsinfodialog import CommandsInfoDialog
import os
import re
from datetime import datetime, timedelta
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
from pygame import mixer
from gtts import gTTS
import tempfile
import time



class SpeechApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setupAudio()
        self.initUI()
        self.listen_thread = ListenThread(self.recognizer, self.mic, 5)
        self.listen_thread.received_audio.connect(self.on_audio_received)
        self.listen_thread.finished.connect(self.on_listen_thread_finished)
        self.listen_thread.debug.connect(self.debug)
        self.setupAutoClose(5)

    def setupAutoClose(self, mn):
        QTimer.singleShot(mn * 60 * 1000, self.closeApp)
        
    def closeApp(self):
        self.close()
        
    def setupAudio(self):
        self.recognizer = sr.Recognizer()
        self.mic = sr.Microphone(0)
        
        mixer.init()
        with self.mic as source:
            self.recognizer.adjust_for_ambient_noise(source,2)
        self.audio_fragments = []
        self.timer_thread = None
        self.listening = False
        self.analyzing = False
        self.stoping = False

    def initUI(self):
        self.setGeometry(300, 300, 200, 200)
        self.setWindowTitle('SpeechApp')
        self.circleButton = CircleButton(self)
        self.circleButton.move(50, 50) 
        self.circleButton.pressed.connect(self.start_listening)
        self.circleButton.released.connect(self.end_listening)
        
        self.infoButton = QPushButton('i', self)
        self.infoButton.clicked.connect(self.showCommandsInfo)
        
    def showCommandsInfo(self):
        dialog = CommandsInfoDialog(self)
        dialog.show()
        
    def resizeEvent(self, event):
        btn_size = int(min(self.width(), self.height()) * 0.9 )  
        x_pos = (self.width() - btn_size ) // 2
        y_pos = (self.height() - btn_size ) // 2
        self.circleButton.setGeometry(x_pos, y_pos, btn_size, btn_size)
        
        self.infoButton.setGeometry(self.width() - int(self.width() / 8),0, int(self.width() / 8),int(self.height() /8) ) 
        super().resizeEvent(event)
    
    def start_listening(self):
        if not self.listen_thread.isListening()  and not self.analyzing:
         
            self.circleButton.setState("listening")
            if not self.listen_thread.isRunning():
                self.listen_thread.start()
                
    def on_audio_received(self, audio):
        #print("reseived")
        self.audio_fragments.append(audio)
    
    def debug(self):
        #print("listen")
        pass     
         
    def end_listening(self):
        self.circleButton.setState("processing")
        self.listen_thread.stop()

    def on_listen_thread_finished(self):
        #print("finished")
        self.listen_thread.stop()
        self.processAudioFragments()
        
    def processAudioFragments(self):
        self.analyzing = True
        if self.audio_fragments:
            combined_audio = sr.AudioData(
                b"".join(fragment.get_raw_data() for fragment in self.audio_fragments),
                self.mic.SAMPLE_RATE,
                self.mic.SAMPLE_WIDTH
            )
      
            self.analyze(combined_audio)
        else:
            self.finishProcessing()
            
    def finishProcessing(self):
        if not self.audio_fragments:
            self.text_to_speech("Нічого не записано.")
        self.audio_fragments.clear()
        self.circleButton.setState("waiting")
        self.analyzing = False   
          
    def analyze(self, audio):
        self.speech_to_text_thread = SpeechToText(self.recognizer, audio)
        self.speech_to_text_thread.result.connect(self.analyze_text)
        self.speech_to_text_thread.error.connect(self.handle_analysis_error)
        self.speech_to_text_thread.start()
    #"один" 
    def analyze_text(self, text):
        print(text)
        commands = [
            (r"(запусти|запуск|постав)(?: таймер|?: таймера)?(?: на)? (\d+)(?: хвилин)?", self.handle_start_timer),
            (r"(почни|почати) запис(?: оцінок|?: записування оцінок)?(?: групи)? (.+)", self.handle_create_excel_file),
             (r"(запиши|записати|додай)(?: оцінку)? (\w+) (\d+)", self.handle_record_grade),
             (r"(відкрий|відкрити) ексель (?: файл)?", self.handle_create_excel_file),
             (r"(відкрий|відкрити)(?: файл)? (.+)", self.handle_open_file),
             (r"скільки(?: часу)?(?: залишилося)?", self.handle_time_left),
             (r"зупини(?:ти)? таймер", self.handle_time_stop)
        ]
        for pattern, handler in commands:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    handler(*match.groups())
                    self.finishProcessing()
                    return 

        self.text_to_speech("Команда не розпізнана.")
        self.finishProcessing()
    
    def handle_start_timer(self, *args):
        minutes = int(args[-1])
        self.start_timer(minutes)

    def handle_record_grade(self, action, surname, grade_str):
        grade = int(grade_str)
        self.record_grade(surname, grade)

    def handle_create_excel_file(self, *args):
        group_name = args[-1]
        self.create_excel_file(group_name)
        
    def handle_open_excel_file(self, *args):
        self.open_file(self.excel_file_path)
        
    def handle_open_file(self, *args):
        file_name = args[-1]
        self.open_file(file_name)

    def handle_time_left(self, *args):
        if self.timer_thread is None or not self.timer_thread.isRunning():
            self.text_to_speech("таймер не запущений")
        else:       
            time_left = self.get_time_left()
            minutes, seconds = divmod(time_left, 60)
            self.text_to_speech(f"Залишилося {minutes} хвилин {seconds} секунд") 
             
    def handle_time_stop(self):
        if self.timer_thread is None or not self.timer_thread.isRunning():
            self.text_to_speech("таймер не запущений")
        else:
            self.timer_thread.stop()
            self.text_to_speech("таймер вимкнений")
    
    def create_excel_file(self, group_name):
        current_date = datetime.now().strftime("%Y-%m-%d") 
        self.excel_file_path = f"{group_name}_{current_date}.xlsx"
        if not os.path.exists(self.excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Прізвище'
            ws['B1'] = 'Оцінка'
            wb.save(self.excel_file_path)
            self.text_to_speech(f"Файл для групи {group_name} створено.")
            
        else:
            self.text_to_speech(f"Файл для групи {group_name} вже існує.")
            
            
    def open_file(self, file_name):
        try:
            os.startfile(file_name) 
        except Exception as e:
            self.text_to_speech(f"Помилка при відкритті файла: {e}")
            
    def record_grade(self, surname, grade):
        
        if not os.path.exists(self.excel_file_path):
            self.text_to_speech("Файл групи не створено")
            return
        else:
            wb = load_workbook(self.excel_file_path)
            ws = wb.active
        max_row = ws.max_row + 1
        ws[f'A{max_row}'] = surname
        ws[f'B{max_row}'] = grade
        wb.save(self.excel_file_path)
        self.text_to_speech(f"Оцінка {grade} для {surname} записана.")
                   
    def handle_analysis_error(self, error):
        self.text_to_speech("Команда не розпізнана.")
        self.finishProcessing()

    def text_to_speech(self, text, lang='uk'):
        tts = gTTS(text=text, lang=lang, slow=False)
        tmpfile_path = tempfile.mktemp(suffix='.mp3')
        tts.save(tmpfile_path)
        mixer.music.load(tmpfile_path)
        mixer.music.play()
        while mixer.music.get_busy():
            time.sleep(0.1)
        mixer.music.unload()
        os.remove(tmpfile_path)
        

    def handle_text_to_speech_error(self, error):
        self.text_to_speech("Помилка обробки тексту.")
        self.circleButton.setState("waiting")
    
    def start_timer(self,minutes):
        if self.timer_thread is not None and self.timer_thread.isRunning():
            self.text_to_speech("Попередній таймер ще не завершено.")
            return
        end_time = datetime.now() + timedelta(minutes=minutes)
        formatted_end_time = end_time.strftime("%H:%M")
        
        self.timer_thread = TimerThread(minutes)
        self.timer_thread.time_left_signal.connect(self.update_time_left)
        self.timer_thread.finished_signal.connect(self.timer_finished)

        self.text_to_speech(f"Таймер запущено на {minutes}. Час завершення {formatted_end_time} хвилин")
        
        self.timer_thread.start()
    
    def update_time_left(self, seconds_left):
        #print(f"Залишилося часу: {seconds_left} секунд")
        pass
    
    def get_time_left(self):
        if self.timer_thread is not None:
            return self.timer_thread.get_time_left()
        return 0
    
    def timer_finished(self):
        self.text_to_speech("Час вийшов")
        
    def closeEvent(self, event):
        if self.timer_thread is not None:
            self.timer_thread.stop()  # Зупинимо поток при закритті програми
        event.accept()
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = SpeechApp()
    ex.show()
    sys.exit(app.exec())
